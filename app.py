import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- STYLES FOR EXCEL GENERATION ---
def apply_excel_theme(ws, title_text):
    """Applies a clean, professional corporate layout to the generated sheet."""
    # Colors & Fonts
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", size=16, bold=True, color="1F4E78")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Title Block
    ws['A1'] = title_text
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30
    
    # Format Headers (Row 3)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 25

    # Format Data Rows
    for row in range(4, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        is_even = (row % 2 == 0)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = regular_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            # Auto-align numbers right, text left
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit column widths dynamically
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# --- ENGINE: EXCEL GENERATION ---
def generate_automated_excel(data_list, doc_types):
    output = io.BytesIO()
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    df = pd.DataFrame(data_list)
    
    # Generate requested tabs
    if "Certificate (Attachment-Ga)" in doc_types:
        ws = wb.create_sheet(title="Attachment-Ga")
        # Structure tailored for official certifications
        ws.append(["Certificate Layout (Attachment-Ga)"]) 
        ws.append([]) # Spacer
        ws.append(["Record ID", "Client/Entity Name", "Reference Code", "Issue Date", "Status"])
        for _, row in df.iterrows():
            ws.append([row.get("ID"), row.get("Name"), f"REF-{row.get('ID')}", row.get("Date"), "VERIFIED"])
        apply_excel_theme(ws, "CERTIFICATE ATTACHMENT (GA)")

    if "Top Sheet SES" in doc_types:
        ws = wb.create_sheet(title="Top Sheet SES")
        ws.append(["Executive Summary - SES"])
        ws.append([])
        ws.append(["Project/Task ID", "Assigned Representative", "Operational Date", "Metric Valuation"])
        for _, row in df.iterrows():
            ws.append([row.get("ID"), row.get("Name"), row.get("Date"), row.get("Value")])
        apply_excel_theme(ws, "TOP SHEET (SES) OVERVIEW")

    if "Top Sheet NMEA" in doc_types:
        ws = wb.create_sheet(title="Top Sheet NMEA")
        ws.append(["Marine/Navigation Summary (NMEA)"])
        ws.append([])
        ws.append(["Log ID", "Vessel/Authority", "Timestamp", "Logged Metrics"])
        for _, row in df.iterrows():
            ws.append([row.get("ID"), row.get("Name"), row.get("Date"), row.get("Value")])
        apply_excel_theme(ws, "TOP SHEET (NMEA) RECONCILIATION")

    if "Detail Sheet (Attachment-Gha)" in doc_types:
        ws = wb.create_sheet(title="Attachment-Gha")
        ws.append(["Granular Itemized Breakdown"])
        ws.append([])
        ws.append(["System Index", "Primary Description", "Log Entry Date", "Base Figure", "Calculated Tax (15%)", "Total Compounded"])
        for _, row in df.iterrows():
            val = float(row.get("Value", 0))
            ws.append([row.get("ID"), row.get("Name"), row.get("Date"), val, val * 0.15, val * 1.15])
        apply_excel_theme(ws, "DETAILED BREAKDOWN SHEET (GHA)")

    wb.save(output)
    return output.getvalue()

# --- USER INTERFACE (STREAMLIT UI) ---
st.set_page_config(page_title="Automated Excel Engine", page_icon="📊", layout="wide")

st.title("📊 Document Generation & Automation Suite")
st.caption("Instantly compile, clean, format, and export custom enterprise sheets.")

st.sidebar.header("⚙️ Configuration Settings")

# 1. Selection Mechanism
document_options = [
    "Certificate (Attachment-Ga)", 
    "Top Sheet SES", 
    "Top Sheet NMEA", 
    "Detail Sheet (Attachment-Gha)"
]
selected_docs = st.sidebar.multiselect(
    "Target Documents to Generate:",
    options=document_options,
    default=document_options[:2]
)

# 2. Input Mode Switcher
input_method = st.radio("Choose Data Input Method:", ("✨ Single Record Form Entry", "📂 Bulk CSV/Excel File Upload"))

final_data = []

# --- MODE A: MANUAL FORM ENTRY ---
if input_method == "✨ Single Record Form Entry":
    st.subheader("📝 Manual Entry Ledger")
    with st.form("manual_entry_form", clear_on_submit=False):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            res_id = st.text_input("Record ID / Index:", "101")
        with col2:
            res_name = st.text_input("Name / Primary Entity:", "Global Corp Ltd")
        with col3:
            res_date = st.date_input("Target Registration Date:")
        with col4:
            res_val = st.number_input("Numerical Assessment Value:", value=25000.00, step=500.0)
            
        submitted = st.form_submit_form("Save & Cache Entry")
        
        # Keep tracking variables saved within the web memory session
        if "data_cache" not in st.session_state:
            st.session_state.data_cache = []
            
        if submitted:
            st.session_state.data_cache.append({
                "ID": res_id, "Name": res_name, "Date": str(res_date), "Value": res_val
            })
            st.success(f"Cached record {res_id} successfully!")
            
    final_data = st.session_state.data_cache

    if final_data:
        st.markdown("### Active Staging Area")
        st.dataframe(pd.DataFrame(final_data), use_container_width=True)
        if st.button("🧹 Clear Staging Memory"):
            st.session_state.data_cache = []
            st.rerun()

# --- MODE B: BULK UPLOAD ---
else:
    st.subheader("📂 Bulk Source Processing")
    uploaded_file = st.file_uploader("Upload reference raw file (CSV or Excel)", type=["csv", "xlsx"])
    
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                uploaded_df = pd.read_csv(uploaded_file)
            else:
                uploaded_df = pd.read_excel(uploaded_file)
                
            st.markdown("### Raw Extracted Stream")
            st.dataframe(uploaded_df.head(10), use_container_width=True)
            
            # Map columns cleanly
            st.info("Mapping dataset columns to generation variables...")
            col_mapping = {}
            cols = list(uploaded_df.columns)
            
            # Smart defaults, but user modifiable mapping dropdowns
            col_mapping["ID"] = st.selectbox("Map unique 'ID' field to:", cols, index=0)
            col_mapping["Name"] = st.selectbox("Map 'Name/Description' field to:", cols, index=min(1, len(cols)-1))
            col_mapping["Date"] = st.selectbox("Map 'Date' field to:", cols, index=min(2, len(cols)-1))
            col_mapping["Value"] = st.selectbox("Map 'Values/Amounts' field to:", cols, index=min(3, len(cols)-1))
            
            # Transform to system standardized dict layout
            transformed_df = uploaded_df.rename(columns={v: k for k, v in col_mapping.items()})
            final_data = transformed_df[["ID", "Name", "Date", "Value"]].to_dict(orient="records")
            
        except Exception as e:
            st.error(f"Error parsing source file: {e}")

# --- PRODUCTION DOWNLOAD HUB ---
st.markdown("---")
st.subheader("🚀 Automation Output Hub")

if not selected_docs:
    st.warning("⚠️ Please select at least one document framework from the sidebar checklist.")
elif not final_data:
    st.info("💡 Ready and listening. Input or upload structural metrics above to activate compilation.")
else:
    st.success(f"System loaded with **{len(final_data)}** items. Compilation engine operational.")
    
    with st.spinner("Executing OpenPyXL matrix layouts..."):
        excel_binary = generate_automated_excel(final_data, selected_docs)
        
    st.download_button(
        label="📥 Download Structured Excel Spreadsheet",
        data=excel_binary,
        file_name="Automated_Report_Suite.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
